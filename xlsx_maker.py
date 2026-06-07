import os
import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

_LOGOS_DIR = os.path.join(os.path.dirname(__file__), "Templates", "Logos")

# ==========================================
# 1. DATABASE PROFIL PERUSAHAAN
# ==========================================
PROFIL_PERUSAHAAN_XLSX = {
    "Angkasa Bangunan": {
        "logo_kiri": os.path.join(_LOGOS_DIR, "angkasa bangunan.png"),
        "nama": "ANGKASA BANGUNAN",
        "tagline": "Specialist Aluminium & Kaca",
        "alamat": "Jl. Contoh Alamat No. 123, Jakarta Barat",
        "header_fill": "D9D9D9",
        "header_font_color": "000000",
    },
    "Angkasa Bangunan Jakarta": {
        "logo_kiri": os.path.join(_LOGOS_DIR, "angkasa bangunan.png"),
        "nama": "PT. ANGKASA BANGUNAN JAKARTA",
        "tagline": "Specialist Aluminium & Kaca",
        "alamat": "Jl. Panjang No. 8, Jakarta Barat",
        "header_fill": "D9D9D9",
        "header_font_color": "000000",
    },
    "Cahaya Kaca Kreatif": {
        "logo_kiri": os.path.join(_LOGOS_DIR, "ckk logo.png"),
        "nama": "PT. CAHAYA KACA KREATIF",
        "tagline": "Applicator Astral Aluminium Systems",
        "alamat": "Jl. Pegangsaan Dua, Kelapa Gading, Jakarta Utara",
        "header_fill": "595959",
        "header_font_color": "FFFFFF",
    },
    "ERI": {
        "logo_kiri": "",
        "nama": "ERI ALUMINIUM",
        "tagline": "Premium Windows & Doors",
        "alamat": "Jl. Sudirman No. 1, Jakarta Pusat",
        "header_fill": "D9D9D9",
        "header_font_color": "000000",
    },
}


# ==========================================
# 2. HELPER STYLES
# ==========================================
def _thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)


def _apply_merged_border(ws, cell_range):
    """Apply thin border to all outer edges of a merged-cell range."""
    thin = Side(style='thin')
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            left   = thin if col == min_col else Side(style=None)
            right  = thin if col == max_col else Side(style=None)
            top    = thin if row == min_row else Side(style=None)
            bottom = thin if row == max_row else Side(style=None)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


# ==========================================
# 3. MESIN UTAMA GENERATOR XLSX
# ==========================================
def generate_quotation_xlsx(keranjang, profit_analysis, client_data):
    profil = PROFIL_PERUSAHAAN_XLSX.get(
        client_data["perusahaan"],
        PROFIL_PERUSAHAAN_XLSX["Angkasa Bangunan"]
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Penawaran Harga"

    # --- COLUMN WIDTHS (matching template) ---
    ws.column_dimensions['A'].width = 7.71
    ws.column_dimensions['B'].width = 20.57
    ws.column_dimensions['C'].width = 41.0
    ws.column_dimensions['D'].width = 5.57
    ws.column_dimensions['E'].width = 23.57
    ws.column_dimensions['F'].width = 25.43
    ws.column_dimensions['G'].width = 8.86

    # --- ROW HEIGHTS FOR HEADER AREA ---
    for i in range(1, 9):
        ws.row_dimensions[i].height = 13.15
    ws.row_dimensions[7].height = 5.1

    # --- COMPANY NAME & TAGLINE (text in cells B1-B3) ---
    ws['B1'] = profil["nama"]
    ws['B1'].font = Font(name='Calibri', bold=True, size=14)
    ws['B1'].alignment = Alignment(vertical='center')

    ws['B2'] = profil["tagline"]
    ws['B2'].font = Font(name='Calibri', size=10)

    ws['B3'] = profil["alamat"]
    ws['B3'].font = Font(name='Calibri', size=9)

    # --- LOGOS ---
    logo_path = profil.get("logo_kiri", "")
    if logo_path and os.path.exists(logo_path):
        img_left = XLImage(logo_path)
        img_left.width = 60
        img_left.height = 60
        ws.add_image(img_left, 'A1')

    astral_path = os.path.join(_LOGOS_DIR, "astral.png")
    if os.path.exists(astral_path):
        img_right = XLImage(astral_path)
        img_right.width = 60
        img_right.height = 60
        ws.add_image(img_right, 'F1')

    # --- DIVIDER LINE (row 7 bottom border) ---
    for col_letter in 'ABCDEF':
        ws[f'{col_letter}7'].border = Border(bottom=Side(style='thin', color='000000'))

    # --- CLIENT INFO SECTION ---
    tgl = datetime.datetime.now().strftime("%d %B %Y")

    ws['A9'] = 'Kepada Yth.'
    ws['A9'].font = Font(name='Calibri', bold=True, size=10)

    ws['F9'] = f'Jakarta, {tgl}'
    ws['F9'].font = Font(name='Calibri', size=10)
    ws['F9'].alignment = Alignment(horizontal='right')

    ws['A10'] = client_data['nama']
    ws['A10'].font = Font(name='Calibri', size=10)

    ws['F10'] = client_data['no_quo']
    ws['F10'].font = Font(name='Calibri', size=10)
    ws['F10'].alignment = Alignment(horizontal='right')

    ws['A11'] = client_data['lokasi']
    ws['A11'].font = Font(name='Calibri', size=10)

    ws['A13'] = (
        'Dengan ini kami sertakan penawaran produk kami sesuai dengan '
        'gambar kerja yang dikirimkan kepada kami.'
    )
    ws['A13'].font = Font(name='Calibri', size=10)
    ws['A13'].alignment = Alignment(wrap_text=True)

    # --- TABLE HEADER (row 16) ---
    h_fill = PatternFill(fill_type='solid', fgColor=profil["header_fill"])
    h_font = Font(name='Calibri', bold=True, size=10, color=profil["header_font_color"])

    for col_letter, label in zip('ABCDEF', ["No", "Item", "Spesifikasi", "Unit", "Harga /Unit", "Harga"]):
        cell = ws[f'{col_letter}16']
        cell.value = label
        cell.font = h_font
        cell.fill = h_fill
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[16].height = 15

    # --- ITEMS SECTION (starting row 17) ---
    CURRENCY = '"Rp "#,##0'
    border_all = _thin_border()

    current_row = 17
    first_item_row = 17

    for i, item in enumerate(keranjang):
        spek_raw = item["meta"].get("spek_custom", "")
        spek_lines = [ln for ln in spek_raw.split("\n") if ln.strip() != ""]

        if not spek_lines:
            spek_lines = ["-"]

        item_start_row = current_row

        for j, line in enumerate(spek_lines):
            r = current_row + j

            ws.row_dimensions[r].height = 13.15

            if j == 0:
                # Main item row: fill all 6 columns
                ws[f'A{r}'] = i + 1
                ws[f'A{r}'].font = Font(name='Calibri', size=10)
                ws[f'A{r}'].border = border_all
                ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='top')

                ws[f'B{r}'] = item["meta"]["nama_item"]
                ws[f'B{r}'].font = Font(name='Calibri', size=10)
                ws[f'B{r}'].border = border_all
                ws[f'B{r}'].alignment = Alignment(horizontal='center', vertical='top')

                ws[f'C{r}'] = line
                ws[f'C{r}'].font = Font(name='Calibri', size=10)
                ws[f'C{r}'].border = border_all
                ws[f'C{r}'].alignment = Alignment(wrap_text=True, vertical='top')

                ws[f'D{r}'] = int(item["meta"]["quantity"])
                ws[f'D{r}'].font = Font(name='Calibri', size=10)
                ws[f'D{r}'].border = border_all
                ws[f'D{r}'].alignment = Alignment(horizontal='center', vertical='top')

                ws[f'E{r}'] = float(item["selling"]["unit_price"])
                ws[f'E{r}'].font = Font(name='Calibri', size=10)
                ws[f'E{r}'].border = border_all
                ws[f'E{r}'].number_format = CURRENCY
                ws[f'E{r}'].alignment = Alignment(horizontal='right', vertical='top')

                ws[f'F{r}'] = f'=D{item_start_row}*E{item_start_row}'
                ws[f'F{r}'].font = Font(name='Calibri', size=10)
                ws[f'F{r}'].border = border_all
                ws[f'F{r}'].number_format = CURRENCY
                ws[f'F{r}'].alignment = Alignment(horizontal='right', vertical='top')

            else:
                # Spec detail rows: only column C has content, others get borders
                for col_letter in 'ABDEF':
                    ws[f'{col_letter}{r}'].border = border_all
                ws[f'C{r}'] = line
                ws[f'C{r}'].font = Font(name='Calibri', size=10)
                ws[f'C{r}'].border = border_all
                ws[f'C{r}'].alignment = Alignment(wrap_text=True, vertical='top')

        current_row += len(spek_lines) + 1  # +1 blank spacer between items

    last_item_row = current_row - 2  # last data row (before the trailing spacer)

    # --- TOTALS SECTION ---
    # Leave one blank row gap above totals
    current_row += 1
    subtotal_row = current_row
    diskon_row   = current_row + 1
    total_row    = current_row + 2

    t_fill = PatternFill(fill_type='solid', fgColor=profil["header_fill"])
    t_font = Font(name='Calibri', bold=True, size=10, color=profil["header_font_color"])

    diskon_pct = client_data['diskon_persen']

    for (row_num, label, formula) in [
        (subtotal_row, "SUB TOTAL",           f'=SUM(F{first_item_row}:F{last_item_row})'),
        (diskon_row,   f'DISKON {diskon_pct}%', f'=F{subtotal_row}*{diskon_pct}%'),
        (total_row,    "TOTAL SETELAH DISKON", f'=F{subtotal_row}-F{diskon_row}'),
    ]:
        merge_range = f'C{row_num}:E{row_num}'
        ws.merge_cells(merge_range)
        _apply_merged_border(ws, merge_range)

        label_cell = ws[f'C{row_num}']
        label_cell.value = label
        label_cell.font = t_font
        label_cell.fill = t_fill
        label_cell.alignment = Alignment(horizontal='center', vertical='center')

        value_cell = ws[f'F{row_num}']
        value_cell.value = formula
        value_cell.font = t_font
        value_cell.fill = t_fill
        value_cell.border = _thin_border()
        value_cell.number_format = CURRENCY
        value_cell.alignment = Alignment(horizontal='right', vertical='center')

        ws.row_dimensions[row_num].height = 15

    # --- NOTES SECTION ---
    notes_start = total_row + 3
    ws[f'A{notes_start}'] = 'Notes :'
    ws[f'A{notes_start}'].font = Font(name='Calibri', bold=True, size=10)

    note_lines = [ln for ln in client_data['note'].split('\n') if ln.strip() != ""]
    for k, line in enumerate(note_lines):
        row_k = notes_start + 1 + k
        ws[f'B{row_k}'] = line
        ws[f'B{row_k}'].font = Font(name='Calibri', size=9)
        ws[f'B{row_k}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_k].height = 13.15

    # --- SIGNATURE SECTION ---
    sig_start = notes_start + len(note_lines) + 3
    sig_end   = sig_start + 6

    ws[f'B{sig_start}'] = 'Disetujui Oleh'
    ws[f'B{sig_start}'].font = Font(name='Calibri', size=10)
    ws[f'B{sig_start}'].alignment = Alignment(horizontal='center')

    ws[f'F{sig_start}'] = 'Hormat Kami'
    ws[f'F{sig_start}'].font = Font(name='Calibri', size=10)
    ws[f'F{sig_start}'].alignment = Alignment(horizontal='center')

    ws[f'B{sig_end}'] = '(                                  )'
    ws[f'B{sig_end}'].font = Font(name='Calibri', size=10)
    ws[f'B{sig_end}'].alignment = Alignment(horizontal='center')

    ws[f'F{sig_end}'] = profil['nama']
    ws[f'F{sig_end}'].font = Font(name='Calibri', size=10)
    ws[f'F{sig_end}'].alignment = Alignment(horizontal='center')

    # --- RETURN AS BYTES ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
